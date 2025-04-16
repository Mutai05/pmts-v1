from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Project, ProjectFeedback, ProjectPhoto, ProjectProgressPhoto, StaffReply, ProjectFeedbackAttachment, StaffReplyAttachment
from .forms import ProjectForm, ProjectProgressForm, ProjectFeedbackForm, ProjectPhotoForm, StaffReplyForm
from locations.models import Ward, SubCounty
from departments.models import Department
from django.http import HttpResponseRedirect, HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime, date
from .filters import FeedbackFilter
from dashboard.views import get_current_term_info # Reuse helper from dashboard app
from django.contrib.auth import get_user_model

User = get_user_model()

def project_list(request):
    """
    Displays a list of all projects with comprehensive filtering, search, and pagination.
    Accessible to all users.
    """
    # --- Get filter parameters --- #
    department_id = request.GET.get('department')
    subcounty_id = request.GET.get('subcounty')
    ward_id = request.GET.get('ward')
    status = request.GET.get('status')
    financial_year = request.GET.get('financial_year')
    search_query = request.GET.get('q')
    project_type = request.GET.get('project_type')
    is_flagship = request.GET.get('is_flagship')
    min_budget = request.GET.get('min_budget')
    max_budget = request.GET.get('max_budget')
    contractor = request.GET.get('contractor')
    min_completion = request.GET.get('min_completion')
    max_completion = request.GET.get('max_completion')
    per_page = request.GET.get('per_page', '20')  # Default to 20 items per page

    # --- Sorting --- #
    sort_by = request.GET.get('sort', '-created_at') # Default sort
    direction = request.GET.get('dir', 'desc')

    # Define valid sortable fields and map them to model fields/annotations
    valid_sort_fields = {
        'name': 'name',
        'department': 'department__name',
        'location': 'ward__subcounty__name', # Sort by SubCounty name first
        'status': 'status',
        'budget': 'budget_allocation',
        'completion': 'percentage_complete',
        'created': 'created_at',
        'end_date': 'end_date',
        # Add more fields as needed
    }

    sort_field = valid_sort_fields.get(sort_by, 'created_at') # Fallback to created_at if invalid
    if direction == 'asc':
        order_by_field = sort_field
    else:
        order_by_field = f'-{sort_field}'

    # Validate per_page
    if per_page not in ['10', '20', '30']:
        per_page = '20'

    # --- Base Queryset --- #
    project_list_qs = Project.objects.select_related(
        'department', 'ward', 'ward__subcounty'
    ).all()

    # --- Apply Filters --- #
    if department_id:
        project_list_qs = project_list_qs.filter(department_id=department_id)
    if subcounty_id:
        project_list_qs = project_list_qs.filter(ward__subcounty_id=subcounty_id)
    if ward_id:
        project_list_qs = project_list_qs.filter(ward_id=ward_id)
    if status:
        project_list_qs = project_list_qs.filter(status=status)
    if financial_year:
        project_list_qs = project_list_qs.filter(financial_year__iexact=financial_year)
    if project_type:
        project_list_qs = project_list_qs.filter(project_type=project_type)
    if is_flagship is not None:
        if is_flagship.lower() in ['true', '1', 'yes']:
            project_list_qs = project_list_qs.filter(is_flagship=True)
        elif is_flagship.lower() in ['false', '0', 'no']:
            project_list_qs = project_list_qs.filter(is_flagship=False)
    if contractor:
        # Assuming contractor is stored in a field named 'contractor' on the Project model
        project_list_qs = project_list_qs.filter(contractor__icontains=contractor)

    if search_query:
        project_list_qs = project_list_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            # Add other searchable fields if needed, e.g.:
            Q(ward__name__icontains=search_query) |
            Q(ward__subcounty__name__icontains=search_query)
            # Q(implementing_agency__icontains=search_query)
        )

    try:
        if min_budget:
            project_list_qs = project_list_qs.filter(budget_allocation__gte=min_budget)
        if max_budget:
            project_list_qs = project_list_qs.filter(budget_allocation__lte=max_budget)
    except (ValueError, TypeError):
        messages.warning(request, "Invalid budget value entered. Filter ignored.")

    try:
        if min_completion:
            project_list_qs = project_list_qs.filter(percentage_complete__gte=min_completion)
        if max_completion:
            project_list_qs = project_list_qs.filter(percentage_complete__lte=max_completion)
    except (ValueError, TypeError):
        messages.warning(request, "Invalid completion percentage value entered. Filter ignored.")

    # Apply Sorting (override default order_by)
    project_list_qs = project_list_qs.order_by(order_by_field)

    # --- Fetch data for filter dropdowns --- #
    departments = Department.objects.all().order_by('name')
    subcounties = SubCounty.objects.all().order_by('name')
    # Wards depend on selected subcounty for potential dynamic filtering (if implemented in frontend JS)
    wards = Ward.objects.all().order_by('name')
    if subcounty_id:
         wards = wards.filter(subcounty_id=subcounty_id)

    status_choices = Project.STATUS_CHOICES
    project_type_choices = Project.PROJECT_TYPE_CHOICES
    flagship_choices = [(None, 'All'), (True, 'Yes'), (False, 'No')]

    # --- Timeframe Filter Logic (Similar to Executive Dashboard) --- #
    all_financial_years = Project.objects.exclude(financial_year__isnull=True).exclude(financial_year='')\
                                        .values_list('financial_year', flat=True).distinct().order_by('-financial_year')
    term_start_year, current_term_value, current_term_label, current_term_start_fy = get_current_term_info()

    timeframe_choices = [
        ('all_time', 'All Time'),
        (current_term_value, current_term_label),
    ]
    timeframe_choices.extend([(f'fy_{fy}', f'FY {fy}') for fy in all_financial_years])

    selected_timeframe = request.GET.get('timeframe', current_term_value) # Default to Current Term

    # Apply timeframe filter (replaces the old simple financial_year filter)
    current_timeframe_label_display = "All Time" # Default display label
    if selected_timeframe.startswith('fy_'):
        fy_value = selected_timeframe.split('fy_')[1]
        project_list_qs = project_list_qs.filter(financial_year=fy_value)
        current_timeframe_label_display = f"FY {fy_value}"
        financial_year = fy_value # For context compatibility
    elif selected_timeframe == current_term_value:
        project_list_qs = project_list_qs.filter(financial_year__gte=current_term_start_fy)
        current_timeframe_label_display = current_term_label
        financial_year = None # Clear specific FY
    elif selected_timeframe == 'all_time':
        # No additional year filter needed for 'all_time'
        current_timeframe_label_display = "All Time"
        financial_year = None # Clear specific FY
    # If the default (current_term_value) is used and the GET param wasn't explicitly set,
    # the filter `financial_year__gte=current_term_start_fy` and the correct label
    # should already be applied from the elif block above.

    # --- Pagination --- #
    paginator = Paginator(project_list_qs, int(per_page))
    page = request.GET.get('page')

    try:
        projects = paginator.page(page)
    except PageNotAnInteger:
        projects = paginator.page(1)
    except EmptyPage:
        projects = paginator.page(paginator.num_pages)

    # --- Build Context --- #
    current_filters = {
        'department': department_id,
        'subcounty': subcounty_id,
        'ward': ward_id,
        'status': status,
        'q': search_query,
        'min_budget': min_budget,
        'max_budget': max_budget,
        'contractor': contractor,
        'min_completion': min_completion,
        'max_completion': max_completion,
        'per_page': per_page,
        'timeframe': selected_timeframe,
        'project_type': project_type,
        'is_flagship': is_flagship,
    }

    # --- Prepare Active Filters for Display --- #
    active_filters = []
    filter_mapping = {
        'department': ('Department', Department, 'name'),
        'subcounty': ('Sub-County', SubCounty, 'name'),
        'ward': ('Ward', Ward, 'name'),
        'status': ('Status', None, None),
        'timeframe': ('Timeframe', None, None), # Use label generated earlier
        'project_type': ('Project Type', None, None),
        'is_flagship': ('Flagship', None, None),
        'q': ('Search', None, None),
        'min_budget': ('Min Budget', None, None),
        'max_budget': ('Max Budget', None, None),
        'contractor': ('Contractor', None, None),
        'min_completion': ('Min %', None, None),
        'max_completion': ('Max %', None, None),
    }

    for key, value in current_filters.items():
        # Check if the filter has a value and is a filter we want to display as a badge
        if value and key != 'per_page' and key in filter_mapping:
            label, model, name_field = filter_mapping[key]
            display_value = value # Default display value

            if key == 'timeframe':
                # Use the pre-calculated display label for timeframe
                if value != 'all_time': # Only show if not 'All Time'
                     display_value = current_timeframe_label_display
                else:
                     continue # Don't show badge for 'All Time' timeframe
            elif model and name_field:
                # Fetch display name for ForeignKeys
                try:
                    instance = model.objects.get(pk=value)
                    display_value = getattr(instance, name_field)
                except (model.DoesNotExist, ValueError, TypeError):
                     display_value = f"Invalid {label}" # Handle case where ID is invalid
            elif key == 'status':
                # Get display name for choice fields
                display_value = dict(Project.STATUS_CHOICES).get(value, value)
            elif key == 'project_type':
                display_value = dict(Project.PROJECT_TYPE_CHOICES).get(value, value)
            elif key == 'is_flagship':
                flagship_map = {'true': 'Yes', '1': 'Yes', 'yes': 'Yes', 'false': 'No', '0': 'No', 'no': 'No'}
                display_value = flagship_map.get(str(value).lower(), 'All')
                if display_value == 'All':
                    continue # Don't show badge for 'All'

            active_filters.append({
                'key': key,
                'label': label,
                'value': display_value
            })

    # Add sorting parameters to context for template links
    current_sort = {
        'by': sort_by if sort_by in valid_sort_fields else 'created_at',
        'dir': direction
    }

    # --- Logic for showing Export button --- #
    show_departmental_export = False
    current_filter_department_id = request.GET.get('department') # Get department filter value

    if request.user.is_authenticated:
        if request.user.role == 'executive':
            show_departmental_export = True # Executives can always export
        elif request.user.role == 'departmental' and request.user.department:
            # Show if no department filter is applied OR if the filter matches the user's department
            # Convert user dept id to string for comparison with GET param
            if current_filter_department_id == str(request.user.department.id):
                 show_departmental_export = True
        # 'public' role won't meet these conditions

    context = {
        'projects': projects,
        'departments': departments,
        'subcounties': subcounties,
        'wards': wards,
        'status_choices': status_choices,
        'project_type_choices': project_type_choices,
        'flagship_choices': flagship_choices,
        'current_sort': current_sort,
        'timeframe_choices': timeframe_choices,
        'selected_timeframe': selected_timeframe,
        'current_filters': current_filters,
        'active_filters': active_filters,
        'show_departmental_export': show_departmental_export,
    }
    return render(request, 'projects/project_list.html', context)


def project_detail(request, slug):
    """
    Display details of a specific project.
    """
    project = get_object_or_404(Project, slug=slug)

    # Get project photos
    photos = project.photos.all() # type: ignore

    # Get progress updates
    progress_updates = project.progress_updates.all() # type: ignore

    # Get feedback that is NOT flagged as inappropriate
    feedback_list = project.feedback.filter(flagged_inappropriate=False).prefetch_related('replies', 'replies__user') # type: ignore

    # Feedback form (public)
    feedback_form = ProjectFeedbackForm()

    # Reply form (staff)
    staff_reply_form = StaffReplyForm()

    if request.method == 'POST':
        # Check if it's a public feedback submission (Check for fields unique to public form)
        if 'comment' in request.POST and 'name' in request.POST and 'id_number' in request.POST:
            public_feedback_form = ProjectFeedbackForm(request.POST)
            if public_feedback_form.is_valid():
                feedback = public_feedback_form.save(commit=False)
                feedback.project = project
                feedback.save()
                messages.success(request, 'Thank you for your feedback!')
                return redirect(request.path + '#feedback-' + str(feedback.id)) # Redirect to new feedback
            else:
                messages.error(request, 'There was an error submitting your feedback. Please check the form.')
                # Ensure the form with errors is passed back to the context
                feedback_form = public_feedback_form # Overwrite the empty form

        # Check if it's a staff reply submission
        elif 'reply_text' in request.POST and 'feedback_id' in request.POST and 'action' in request.POST and request.POST['action'] == 'post_reply':
            feedback_id = request.POST.get('feedback_id')
            try:
                feedback_instance = get_object_or_404(ProjectFeedback, id=feedback_id, project=project)

                # Authorization Check:
                is_authorized = False
                if request.user.is_authenticated:
                    if request.user.role == 'executive':
                        is_authorized = True
                    elif request.user.role == 'departmental' and request.user.department == feedback_instance.project.department:
                        is_authorized = True

                if is_authorized:
                    reply_form_posted = StaffReplyForm(request.POST, request.FILES) # Pass request.FILES
                    if reply_form_posted.is_valid():
                        reply = reply_form_posted.save(commit=False)
                        reply.feedback = feedback_instance
                        reply.user = request.user
                        reply.save() # Save reply first to get an ID

                        # Handle multiple file uploads for the reply
                        for f in request.FILES.getlist(f'reply_attachments_{feedback_id}'): # Use feedback_id for unique name
                            is_image = f.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))
                            StaffReplyAttachment.objects.create(
                                reply=reply,
                                file=f,
                                is_image=is_image
                            )

                        messages.success(request, 'Your reply has been posted.')
                        return redirect(request.path + f'#feedback-{feedback_id}')
                    else:
                        messages.error(request, 'There was an error posting your reply. Please check the form below the feedback item.')
                        # Store the specific form with errors to display it below the correct feedback item
                        # We need a way to associate this errored form with the specific feedback ID in the template
                        # Option 1: Pass a dictionary of errored forms in context
                        # Option 2: Use session (more complex)
                        # Let's try Option 1 (though context modification here is tricky)
                        # For now, just passing the last errored form might show it everywhere, need template logic adjustment
                        staff_reply_form = reply_form_posted # Pass back the form with errors
                else:
                    messages.error(request, 'You are not authorized to post replies to this feedback.')

            except ProjectFeedback.DoesNotExist:
                messages.error(request, 'Feedback item not found.')

        # Check if it's a flag/report action
        elif 'action' in request.POST and request.POST.get('action') in ['toggle_flag', 'toggle_report'] and 'feedback_id' in request.POST:
            feedback_id = request.POST.get('feedback_id')
            try:
                feedback_instance = get_object_or_404(ProjectFeedback, id=feedback_id, project=project)
                action = request.POST.get('action')

                if request.user.is_authenticated:
                    # Staff flagging action
                    if action == 'toggle_flag' and request.user.role in ['departmental', 'executive']:
                        # Departmental users can only flag feedback in their department
                        if request.user.role == 'executive' or (request.user.role == 'departmental' and request.user.department == feedback_instance.project.department):
                            feedback_instance.flagged_inappropriate = not feedback_instance.flagged_inappropriate
                            feedback_instance.save(update_fields=['flagged_inappropriate'])
                            status = "flagged" if feedback_instance.flagged_inappropriate else "unflagged"
                            messages.success(request, f'Feedback has been {status}.')
                        else:
                            messages.error(request, 'Departmental users can only flag feedback for their department.')

                    # Public reporting action
                    elif action == 'toggle_report': # Any authenticated user can report
                        feedback_instance.reported_inappropriate = not feedback_instance.reported_inappropriate
                        feedback_instance.save(update_fields=['reported_inappropriate'])
                        status = "reported as inappropriate" if feedback_instance.reported_inappropriate else "report removed"
                        messages.success(request, f'Feedback status updated: {status}.')
                    else:
                        # User role doesn't match the action attempted
                         messages.error(request, 'You do not have permission to perform this action.')
                else:
                     messages.error(request, 'You must be logged in to flag or report feedback.')

                # Redirect after action
                return redirect(request.path + f'#feedback-{feedback_id}')

            except ProjectFeedback.DoesNotExist:
                messages.error(request, 'Feedback item not found.')

    context = {
        'project': project,
        'photos': photos,
        'progress_updates': progress_updates,
        'feedback_list': feedback_list,
        'feedback_form': feedback_form, # Public feedback form
        'staff_reply_form': staff_reply_form, # Staff reply form
    }

    return render(request, 'projects/project_detail.html', context)


@login_required
def project_create(request):
    """
    Create a new project.
    """
    if request.user.role not in ['executive', 'departmental']:
        messages.error(request, 'You do not have permission to create projects.')
        return redirect('dashboard:home')

    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save()
            messages.success(request, f'Project "{project.name}" has been created.')
            return redirect('projects:project_detail', slug=project.slug)
        import pdb; pdb.set_trace()
    else:
        # Pre-select department for departmental users
        initial = {}
        if request.user.role == 'departmental' and request.user.department:
            initial['department'] = request.user.department

        form = ProjectForm(initial=initial)

    context = {
        'form': form,
        'title': 'Create Project'
    }

    return render(request, 'projects/project_form.html', context)


@login_required
def project_update(request, slug):
    """
    Update an existing project.
    """
    if request.user.role not in ['executive', 'departmental']:
        messages.error(request, 'You do not have permission to update projects.')
        return redirect('dashboard:home')

    project = get_object_or_404(Project, slug=slug)

    # Check if user has permission to edit this project
    if request.user.role == 'departmental' and request.user.department != project.department:
        messages.error(request, 'You do not have permission to update this project.')
        return redirect('projects:project_detail', slug=project.slug)

    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            updated_project = form.save()

            # --- Handle Dynamically Added Photos and Captions --- #
            index = 0
            while True:
                photo_field_name = f'new_photo_{index}'
                caption_field_name = f'new_caption_{index}'

                if photo_field_name not in request.FILES:
                    break # Stop when we run out of photo fields

                uploaded_file = request.FILES[photo_field_name]
                caption = request.POST.get(caption_field_name, '') # Get caption, default to empty

                # Check if there's already a cover photo ONLY for the first potential upload (index 0)
                # Otherwise, subsequent photos in this submission won't be cover photos.
                is_cover_photo = False
                if index == 0:
                    has_cover = ProjectPhoto.objects.filter(project=updated_project, is_cover=True).exists()
                    if not has_cover:
                        is_cover_photo = True

                ProjectPhoto.objects.create(
                    project=updated_project,
                    image=uploaded_file,
                    caption=caption,
                    is_cover=is_cover_photo
                )
                index += 1
             # --- End Photo Handling --- #

            messages.success(request, f'Project "{updated_project.name}" has been updated.')
            return redirect('projects:project_detail', slug=updated_project.slug)
    else:
        form = ProjectForm(instance=project)

    context = {
        'form': form,
        'title': 'Update Project',
        'project': project
    }

    return render(request, 'projects/project_form.html', context)


@login_required
def project_progress_update(request, slug):
    """
    Add a progress update to a project, including multiple photos.
    """
    if request.user.role not in ['executive', 'departmental']:
        messages.error(request, 'You do not have permission to update project progress.')
        return redirect('dashboard:home')

    project = get_object_or_404(Project, slug=slug)

    # Check if user has permission to edit this project
    if request.user.role == 'departmental' and request.user.department != project.department:
        messages.error(request, 'You do not have permission to update this project.')
        return redirect('projects:project_detail', slug=project.slug)

    if request.method == 'POST':
        form = ProjectProgressForm(request.POST, request.FILES)
        if form.is_valid():
            progress = form.save(commit=False)
            progress.project = project
            progress.updated_by = request.user
            progress.save() # Save the main progress update first

            # --- Handle Dynamically Added Photos and Captions --- #
            index = 0
            while True:
                photo_field_name = f'progress_photo_{index}'
                caption_field_name = f'progress_caption_{index}'

                if photo_field_name not in request.FILES:
                    break # Stop when we run out of photo fields

                uploaded_file = request.FILES[photo_field_name]
                caption = request.POST.get(caption_field_name, '') # Get caption, default to empty

                ProjectProgressPhoto.objects.create(
                    progress_update=progress,
                    image=uploaded_file,
                    caption=caption
                )
                index += 1
            # --- End Photo Handling --- #

            messages.success(request, f'Progress update added to "{project.name}".')
            return redirect('projects:project_detail', slug=project.slug)
    else:
        form = ProjectProgressForm()

    context = {
        'form': form,
        'project': project,
        'title': 'Add Progress Update'
    }
    return render(request, 'projects/progress_form.html', context)


def project_feedback(request, slug):
    """
    Handles the submission of public feedback for a specific project, including file uploads.
    """
    project = get_object_or_404(Project, slug=slug)
    if request.method == 'POST':
        form = ProjectFeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.project = project
            feedback.save()

            # Handle multiple file uploads
            for f in request.FILES.getlist('attachments'):
                # Basic check for image type based on extension (can be expanded)
                is_image = f.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))
                ProjectFeedbackAttachment.objects.create(
                    feedback=feedback,
                    file=f,
                    is_image=is_image
                    # caption can be added later if needed via formset or separate mechanism
                )

            messages.success(request, 'Your feedback has been submitted successfully. Thank you!')
            return redirect('projects:project_detail', slug=slug) # Redirect back to project page
        else:
            # If form is invalid, re-render the detail page with the form errors
            # We need to fetch the project details again for the detail page context
            messages.error(request, 'Please correct the errors below and resubmit your feedback.')
            # Redirect back to the detail page, potentially passing errors (or let detail view handle it)
            # A simple redirect is easier, the detail view will re-render the form with errors
            # Store form errors in session? Or rely on detail view re-rendering?
            # Let's redirect and let the detail view handle re-rendering the form.
            # This approach loses the invalid form data, which is not ideal UX.
            # TODO: Improve error handling to retain form data on validation failure.
            return redirect('projects:project_detail', slug=slug)

    # If GET request or form submission failed, redirect to the project detail page
    # (which should display the form)
    return redirect('projects:project_detail', slug=slug)


@login_required
def project_manage_photos(request, slug):
    """
    Manage photos for a specific project (upload, delete, set cover).
    """
    project = get_object_or_404(Project, slug=slug)

    # Permission Check: Executive or Departmental user of the correct department
    if not (request.user.role == 'executive' or
            (request.user.role == 'departmental' and request.user.department == project.department)):
        messages.error(request, 'You do not have permission to manage photos for this project.')
        return redirect('projects:project_detail', slug=project.slug)

    # Handle Photo Deletion
    if request.method == 'POST' and 'delete_photo' in request.POST:
        photo_id = request.POST.get('photo_id')
        try:
            photo_to_delete = ProjectPhoto.objects.get(id=photo_id, project=project)
            photo_to_delete.delete()
            messages.success(request, 'Photo deleted successfully.')
        except ProjectPhoto.DoesNotExist:
            messages.error(request, 'Photo not found or you do not have permission to delete it.')
        return HttpResponseRedirect(request.path_info) # Redirect back to the same page

    # Handle Cover Photo Setting
    if request.method == 'POST' and 'set_cover' in request.POST:
        photo_id = request.POST.get('photo_id')
        try:
            photo_to_set = ProjectPhoto.objects.get(id=photo_id, project=project)
            # Unset other cover photos for this project
            project.photos.update(is_cover=False)
            # Set the new cover photo
            photo_to_set.is_cover = True
            photo_to_set.save()
            messages.success(request, 'Cover photo updated successfully.')
        except ProjectPhoto.DoesNotExist:
            messages.error(request, 'Photo not found or you do not have permission to set it as cover.')
        return HttpResponseRedirect(request.path_info) # Redirect back to the same page

    # Handle Photo Upload
    if request.method == 'POST':
        form = ProjectPhotoForm(request.POST, request.FILES)
        if form.is_valid():
            photo = form.save(commit=False)
            photo.project = project

            # Handle the 'is_cover' logic correctly
            if photo.is_cover:
                # If this photo is marked as cover, unset any existing cover photo
                project.photos.update(is_cover=False)
            elif not project.photos.filter(is_cover=True).exists():
                 # If no cover photo exists yet, make this the cover by default
                 # Only if the form didn't explicitly set it to False
                 if 'is_cover' not in request.POST:
                    photo.is_cover = True

            photo.save()
            messages.success(request, 'Photo uploaded successfully.')
            return HttpResponseRedirect(request.path_info) # Redirect to refresh the page
        else:
            # Display form errors if validation fails
            messages.error(request, "Error uploading photo. Please check the form.")
    else:
        form = ProjectPhotoForm()

    # Get existing photos for display
    photos = project.photos.all()

    context = {
        'project': project,
        'photos': photos,
        'form': form,
        'title': f'Manage Photos for {project.name}'
    }
    return render(request, 'projects/project_manage_photos.html', context)


@login_required
def export_projects_excel(request):
    """
    Exports filtered project data to an Excel file.
    Only accessible by executive and departmental staff.
    """
    if request.user.role not in ['executive', 'departmental']:
        messages.error(request, 'You do not have permission to export project data.')
        return redirect('dashboard:home')

    # Reuse filtering logic from project_list view
    projects = Project.objects.all()
    if request.user.role == 'departmental' and request.user.department:
        # Departmental users can only export their department's projects
        projects = projects.filter(department=request.user.department)

    # Get filter parameters from the request (passed from the list page)
    department_id = request.GET.get('department')
    subcounty_id = request.GET.get('subcounty')
    ward_id = request.GET.get('ward')
    status = request.GET.get('status')
    financial_year = request.GET.get('financial_year')
    search_query = request.GET.get('q')
    min_budget = request.GET.get('min_budget')
    max_budget = request.GET.get('max_budget')
    contractor = request.GET.get('contractor')
    min_completion = request.GET.get('min_completion')
    max_completion = request.GET.get('max_completion')
    project_type = request.GET.get('project_type')
    is_flagship = request.GET.get('is_flagship')

    # Apply filters (Executive can filter by any department, departmental is already filtered)
    if request.user.role == 'executive' and department_id:
         projects = projects.filter(department_id=department_id)
    # Note: Departmental users already filtered above, so we don't apply department_id again for them.

    if subcounty_id:
        projects = projects.filter(ward__subcounty_id=subcounty_id)
    if ward_id:
        projects = projects.filter(ward_id=ward_id)
    if status:
        projects = projects.filter(status=status)
    if financial_year:
        projects = projects.filter(financial_year__iexact=financial_year)
    if project_type:
        projects = projects.filter(project_type=project_type)
    if is_flagship is not None:
        if is_flagship.lower() in ['true', '1', 'yes']:
            projects = projects.filter(is_flagship=True)
        elif is_flagship.lower() in ['false', '0', 'no']:
             projects = projects.filter(is_flagship=False)
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(implementing_agency__icontains=search_query) |
            Q(contractor_name__icontains=search_query)
        )
    try:
        if min_budget:
            projects = projects.filter(budget_allocation__gte=min_budget)
        if max_budget:
            projects = projects.filter(budget_allocation__lte=max_budget)
    except (ValueError, TypeError):
        pass # Ignore invalid filter values for export
    if contractor:
        projects = projects.filter(contractor_name__icontains=contractor)
    try:
        if min_completion:
            projects = projects.filter(percentage_complete__gte=min_completion)
        if max_completion:
            projects = projects.filter(percentage_complete__lte=max_completion)
    except (ValueError, TypeError):
        pass # Ignore invalid filter values for export

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects Export"

    # Define Headers
    headers = [
        "Project Name", "Department", "Sub-County", "Ward", "Status", "Project Type", "Flagship",
        "Start Date", "End Date", "% Complete", "Budget Allocation (KES)",
        "Expenditure (KES)", "Budget Utilization (%)", "Funding Source",
        "Contractor", "Contractor Phone", "Contractor Email", "Contractor Address",
        "Implementing Agency", "Challenges", "Google Map Link"
    ]
    ws.append(headers)

    # Style Headers
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        # Auto-adjust column width (approximate)
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = len(header_title) + 5

    # Populate data rows
    row_num = 2
    for project in projects:
        utilization = project.budget_utilization if project.budget_utilization else 0
        row = [
            project.name,
            project.department.name,
            project.subcounty.name,
            project.ward.name,
            project.get_status_display(),
            project.get_project_type_display(),
            "Yes" if project.is_flagship else "No",
            project.start_date,
            project.end_date,
            project.percentage_complete,
            project.budget_allocation,
            project.expenditure,
            round(utilization, 2),
            project.get_funding_source_display(),
            project.contractor_name or "",
            project.contractor_phone or "",
            project.contractor_email or "",
            project.contractor_address or "",
            project.implementing_agency,
            project.challenges or "",
            project.google_map_location or ""
        ]
        ws.append(row)
        # Apply number formatting for currency and percentage columns
        ws.cell(row=row_num, column=10).number_format = '0' # Percentage Complete
        ws.cell(row=row_num, column=11).number_format = '#,##0.00' # Budget
        ws.cell(row=row_num, column=12).number_format = '#,##0.00' # Expenditure
        ws.cell(row=row_num, column=13).number_format = '0.00' # Utilization %
        # Apply date formatting (Excel handles Python date objects well)
        ws.cell(row=row_num, column=8).number_format = 'yyyy-mm-dd' # Start Date
        ws.cell(row=row_num, column=9).number_format = 'yyyy-mm-dd' # End Date
        row_num += 1

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="projects_{timestamp}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


@login_required
def feedback_dashboard(request):
    """
    Dashboard for staff to view, filter, and reply to public feedback.
    """
    today = date.today()
    # Authorization: Only Departmental and Executive users
    if request.user.role not in ['departmental', 'executive']:
        messages.error(request, "You do not have permission to access this page.")
        return redirect('dashboard:home') # Or appropriate redirect

    # Base queryset
    feedback_qs = ProjectFeedback.objects.select_related(
        'project', 'project__department', 'sub_county', 'ward'
    ).prefetch_related('replies', 'replies__user').all()

    # Apply role-based filtering
    if request.user.role == 'departmental':
        feedback_qs = feedback_qs.filter(project__department=request.user.department)
    errored_reply_form = None
    # --- Handle Actions (Flagging, Replying) --- #
    if request.method == 'POST':
        action = request.POST.get('action')
        feedback_id = request.POST.get('feedback_id')

        try:
            feedback_item = ProjectFeedback.objects.get(id=feedback_id)
            # Ensure user has permission for this item (redundant if queryset is filtered, but good practice)
            if request.user.role == 'departmental' and feedback_item.project.department != request.user.department:
                messages.error(request, "You do not have permission to modify this feedback.")
            else:
                if action == 'toggle_flag':
                    feedback_item.flagged_inappropriate = not feedback_item.flagged_inappropriate
                    feedback_item.save(update_fields=['flagged_inappropriate'])
                    status = "flagged" if feedback_item.flagged_inappropriate else "unflagged"
                    messages.success(request, f"Feedback has been {status}.")

                elif action == 'add_reply':
                    reply_form = StaffReplyForm(request.POST, request.FILES) # Pass FILES here
                    if reply_form.is_valid():
                        reply = reply_form.save(commit=False)
                        reply.feedback = feedback_item
                        reply.user = request.user
                        reply.save() # Save reply first

                        # Handle reply attachments
                        attachment_input_name = f'reply_attachments_{feedback_id}'
                        for f in request.FILES.getlist(attachment_input_name):
                            is_image = f.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))
                            StaffReplyAttachment.objects.create(
                                reply=reply,
                                file=f,
                                is_image=is_image
                            )

                        messages.success(request, "Reply added successfully.")
                    else:
                        # Store the errored form to pass back to context
                        errored_reply_form = reply_form
                        # Associate feedback_id with the errored form for template logic
                        errored_reply_form.instance.feedback_id = feedback_id
                        messages.error(request, f"Error submitting reply for feedback ID {feedback_id}. Please check the form below the feedback item.")

        except ProjectFeedback.DoesNotExist:
            messages.error(request, "Feedback item not found.")
        except Exception as e:
             messages.error(request, f"An error occurred: {e}")

        # Redirect back to the same page (with filters potentially)
        # This might lose the exact page number, consider passing query params
        # Or stay on the page if there was a reply form error
        if not errored_reply_form:
            return HttpResponseRedirect(request.path_info + '?' + request.GET.urlencode() + f'#feedback-{feedback_id}')
        # If there was an error, we fall through to render the page again with the errored form in context

    # --- Filtering --- #
    # Default filter: Exclude flagged unless status filter is active
    status_filter = request.GET.get('status')
    if not status_filter or status_filter == 'all':
         # Apply default filter *before* passing to FeedbackFilter if status is 'all' or not set
         initial_qs = feedback_qs.filter(flagged_inappropriate=False, reported_inappropriate=False)
    else:
        initial_qs = feedback_qs # Let the filter handle flagged/reported statuses

    feedback_filter = FeedbackFilter(request.GET, queryset=initial_qs, user=request.user)
    filtered_qs = feedback_filter.qs

    # The filtering for time_range is handled by the FeedbackFilter class method
    # Get the selected time range (defaulting to 'month' if not provided)
    selected_time_range = request.GET.get('time_range', 'month')

    # --- Pagination --- #
    allowed_per_page = [10, 20, 30, 50]
    try:
        per_page = int(request.GET.get('per_page', 30)) # Default to 30
        if per_page not in allowed_per_page:
            per_page = 30 # Fallback to default if invalid value
    except ValueError:
        per_page = 30 # Fallback to default if not an integer

    paginator = Paginator(filtered_qs, per_page)
    page = request.GET.get('page')
    try:
        feedback_page = paginator.page(page)
    except PageNotAnInteger:
        feedback_page = paginator.page(1)
    except EmptyPage:
        feedback_page = paginator.page(paginator.num_pages)

    # Use the potentially errored form, or a clean one
    # We pass the *specific* errored form instance if it exists
    if errored_reply_form:
        reply_form_instance = errored_reply_form
    else:
        reply_form_instance = StaffReplyForm()

    # --- Prepare Active Filters for Display --- #
    active_filters = []
    # Create a mutable copy of GET params to potentially add defaults for badge display
    filter_params_for_badges = request.GET.copy()

    # Ensure the default time_range ('month') is considered for badge display if not specified
    if 'time_range' not in filter_params_for_badges:
        filter_params_for_badges['time_range'] = 'month'

    filter_mapping = {
        'project': ('Project', Project, 'name'),
        'department': ('Department', Department, 'name'),
        'sub_county': ('Sub-County', SubCounty, 'name'),
        'ward': ('Ward', Ward, 'name'),
        'rating': ('Rating', None, None),
        'status': ('Status', None, None),
        'time_range': ('Date Range', None, None),
        'project_type': ('Project Type', None, None),
        'is_flagship': ('Flagship', None, None),
    }

    rating_choices_dict = dict(ProjectFeedback.RATING_CHOICES)
    status_choices_dict = dict(FeedbackFilter.STATUS_CHOICES)
    time_range_choices_dict = dict(FeedbackFilter.TIME_RANGE_CHOICES)

    # Use the modified dictionary to generate badges
    for key, value in filter_params_for_badges.items():
        # Process only if value exists and key is a displayable filter
        if value and key in filter_mapping:
            label, model, name_field = filter_mapping[key]
            display_value = value # Start with the raw value
            should_append = True # Flag to control appending

            if model and name_field:
                try:
                    instance = model.objects.get(pk=value)
                    display_value = getattr(instance, name_field)
                except (model.DoesNotExist, ValueError, TypeError):
                    display_value = f"Invalid {label}"
            elif key == 'rating':
                 try:
                    display_value = rating_choices_dict.get(int(value), value)
                 except (ValueError, TypeError):
                    display_value = value
            elif key == 'status':
                 display_value = status_choices_dict.get(value, value)
            elif key == 'time_range':
                # Always look up the display value from choices
                display_value = time_range_choices_dict.get(value, value)
            elif key == 'project_type':
                display_value = dict(Project.PROJECT_TYPE_CHOICES).get(value, value)
            elif key == 'is_flagship':
                flagship_map = {'true': 'Yes', '1': 'Yes', 'yes': 'Yes', 'false': 'No', '0': 'No', 'no': 'No'}
                display_value = flagship_map.get(str(value).lower(), 'All')
                if display_value == 'All':
                    continue # Don't show badge for 'All'

            # Append the filter badge info if should_append is True
            if should_append:
                active_filters.append({
                   'key': key,
                   'label': label,
                   'value': display_value
                })

    context = {
        'filter': feedback_filter,
        'feedback_page': feedback_page,
        'reply_form': reply_form_instance, # Pass the potentially errored form
        'current_per_page': per_page,
        'allowed_per_page_options': allowed_per_page,
        'active_filters': active_filters,
        'title': 'Feedback Dashboard'
    }
    return render(request, 'projects/feedback_dashboard.html', context)


@login_required
def project_delete(request, slug):
    project = get_object_or_404(Project, slug=slug)
    if request.method == 'POST':
        project.delete()
        messages.success(request, 'Project deleted successfully.')
        return redirect('projects:project_list')
    return render(request, 'projects/project_confirm_delete.html', {'project': project})
